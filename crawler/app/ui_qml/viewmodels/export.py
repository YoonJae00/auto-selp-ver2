from __future__ import annotations

from collections.abc import Callable
from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from PySide6.QtCore import Property, QTimer, QUrl, Signal, Slot
from PySide6.QtWidgets import QFileDialog
from sqlalchemy import func, or_, select

from app.db.models import Product, ProductOption, Supplier
from app.db.session import get_session
from app.paths import exports_dir as default_exports_dir
from app.ui_qml.models.list_model import ListModel
from app.ui_qml.viewmodels.base import BaseViewModel, sanitize_diagnostic
from app.workers.export import ExportRequest, ExportWorker
from app.workers.lifecycle import stop_workers


ISSUE_ROLES = ("severity", "code", "message", "productId", "productCode")


def _load_suppliers() -> list[Supplier]:
    session = get_session()
    try:
        return list(session.scalars(select(Supplier).order_by(Supplier.name)))
    finally:
        session.close()


def _load_scope(supplier_id: str) -> tuple[int, int, list[dict[str, str]]]:
    session = get_session()
    try:
        product_filter = Product.supplier_id == supplier_id
        product_count = int(session.scalar(select(func.count(Product.id)).where(product_filter)) or 0)
        option_count = int(session.scalar(
            select(func.count(ProductOption.id)).join(Product).where(product_filter)
        ) or 0)
        issues: list[dict[str, str]] = []
        if not product_count:
            return 0, option_count, issues
        # Bound representative rows so validation remains responsive for very large scopes.
        candidates = list(session.scalars(select(Product).where(
            product_filter,
            or_(
                Product.raw_product_name == "", Product.supplier_product_code == "",
                Product.supplier_status == "", Product.origin.is_(None),
                Product.supply_price.is_(None), Product.main_image_url.is_(None),
            ),
        ).order_by(Product.supplier_product_code).limit(50)))
        for product in candidates:
            common = {"productId": product.id, "productCode": product.supplier_product_code or ""}
            for field, value, severity, label in (
                ("raw_product_name", product.raw_product_name, "error", "상품명"),
                ("supplier_product_code", product.supplier_product_code, "error", "상품 코드"),
                ("supplier_status", product.supplier_status, "error", "상품 상태"),
                ("origin", product.origin, "warning", "원산지"),
                ("supply_price", product.supply_price, "warning", "공급가"),
                ("main_image_url", product.main_image_url, "warning", "대표 이미지"),
            ):
                if value is None or value == "":
                    issues.append({**common, "severity": severity, "code": f"missing_{field}", "message": f"{label}이(가) 없습니다."})
        issues.sort(key=lambda row: (row["severity"] != "error", row["productCode"], row["code"]))
        if len(issues) > 50 or len(candidates) == 50:
            issues = issues[:49]
            issues.append({"severity": "warning", "code": "more_issues", "message": "추가 누락 항목이 있을 수 있습니다.", "productId": "", "productCode": ""})
        return product_count, option_count, issues
    finally:
        session.close()


class ExportViewModel(BaseViewModel):
    stateChanged = Signal()

    def __init__(self, parent=None, *, app_view_model=None,
                 supplier_loader: Callable[[], list[Any]] = _load_suppliers,
                 scope_loader: Callable[[str], tuple[int, int, list[dict[str, str]]]] = _load_scope,
                 exports_dir: Path | None = None, picker: Callable[[str], Any] | None = None,
                 worker_factory: Callable[[ExportRequest], Any] = ExportWorker) -> None:
        super().__init__(parent)
        self._app = app_view_model
        self._scope_loader = scope_loader
        self._exports_dir = Path(exports_dir) if exports_dir is not None else default_exports_dir()
        self._picker = picker
        self._worker_factory = worker_factory
        self._supplier_ids: set[str] = set()
        self._supplier_names: dict[str, str] = {}
        self._supplier_id = ""
        self._product_count = self._option_count = 0
        self._issues = ListModel(ISSUE_ROLES, parent=self)
        self._history = ListModel(("fileName", "path", "exportedAt", "rowCount", "outcome"), parent=self)
        self._suppliers = ListModel(("id", "name"), parent=self)
        self._warning_acknowledged = False
        self._validated = False
        self._output_path: Path | None = None
        self._busy = False
        self._worker = None
        self._retired_workers: list[Any] = []
        self._operation_id = 0
        self._task_owner: object | None = None
        self._shutting_down = False
        suppliers = supplier_loader()
        rows = [{"id": str(item.id), "name": item.name} for item in suppliers]
        self._supplier_ids = {row["id"] for row in rows}
        self._supplier_names = {row["id"]: row["name"] for row in rows}
        self._suppliers.resetRows(rows)
        self.refreshHistory()

    suppliers = Property(object, lambda self: self._suppliers, constant=True)
    issues = Property(object, lambda self: self._issues, constant=True)
    history = Property(object, lambda self: self._history, constant=True)
    selectedSupplierId = Property(str, lambda self: self._supplier_id, notify=stateChanged)
    productCount = Property(int, lambda self: self._product_count, notify=stateChanged)
    optionCount = Property(int, lambda self: self._option_count, notify=stateChanged)
    warningAcknowledged = Property(bool, lambda self: self._warning_acknowledged, notify=stateChanged)
    busy = Property(bool, lambda self: self._busy, notify=stateChanged)
    destinationName = Property(str, lambda self: self._output_path.name if self._output_path else "", notify=stateChanged)

    @Property(bool, notify=stateChanged)
    def canExport(self) -> bool:
        rows = self._issues._rows
        errors = any(row["severity"] == "error" for row in rows)
        warnings = any(row["severity"] == "warning" for row in rows)
        retired_running = any(getattr(worker, "isRunning", lambda: False)() for worker in self._retired_workers)
        return bool(self._validated and self._supplier_id and self._output_path and not self._busy and not retired_running and not errors and (not warnings or self._warning_acknowledged))

    def _emit(self) -> None:
        self.stateChanged.emit()
        self.changed.emit()

    @Slot(str)
    def setSupplierId(self, supplier_id: str) -> None:
        value = supplier_id if supplier_id in self._supplier_ids else ""
        if value == self._supplier_id or self._busy:
            return
        self._supplier_id = value
        self._product_count = self._option_count = 0
        self._issues.resetRows([])
        self._warning_acknowledged = self._validated = False
        self._output_path = None
        self._emit()

    @Slot(result=bool)
    def validateScope(self) -> bool:
        if not self._supplier_id:
            counts = (0, 0)
            issues = [{"severity": "error", "code": "supplier_required", "message": "도매처를 선택하세요.", "productId": "", "productCode": ""}]
        else:
            try:
                product_count, option_count, issues = self._scope_loader(self._supplier_id)
            except Exception as exc:
                product_count, option_count = 0, 0
                issues = [{"severity": "error", "code": "validation_failed", "message": sanitize_diagnostic(exc), "productId": "", "productCode": ""}]
            counts = (product_count, option_count)
            if product_count == 0:
                issues.insert(0, {"severity": "error", "code": "empty_scope", "message": "내보낼 상품이 없습니다.", "productId": "", "productCode": ""})
        self._product_count, self._option_count = counts
        issues.sort(key=lambda row: row.get("severity") != "error")
        self._issues.resetRows(issues)
        self._warning_acknowledged = False
        self._validated = True
        if self._supplier_id and self._output_path is None:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            supplier_name = re.sub(r"[^\w.-]+", "_", self._supplier_names.get(self._supplier_id, self._supplier_id)).strip("._") or "supplier"
            self._output_path = self._exports_dir / f"{supplier_name}_{stamp}.xlsx"
        self._emit()
        return not any(row["severity"] == "error" for row in issues)

    @Slot(bool)
    def acknowledgeWarnings(self, acknowledged: bool = True) -> None:
        self._warning_acknowledged = bool(acknowledged)
        self._emit()

    @Slot(str)
    def setOutputPath(self, path: str) -> None:
        candidate = Path(path).expanduser()
        if candidate.suffix.lower() != ".xlsx":
            candidate = candidate.with_suffix(".xlsx")
        self._output_path = candidate
        self._emit()

    @Slot()
    def chooseOutputFile(self) -> None:
        initial = str(self._output_path or self._exports_dir / "export.xlsx")
        if self._picker:
            try:
                selected = self._picker(initial)
            except TypeError:
                selected = self._picker()
        else:
            selected = QFileDialog.getSaveFileUrl(None, "Excel 내보내기", QUrl.fromLocalFile(initial), "Excel (*.xlsx)")[0]
        if isinstance(selected, tuple):
            selected = selected[0]
        path = selected.toLocalFile() if isinstance(selected, QUrl) else str(selected or "")
        if path:
            self.setOutputPath(path)

    @Slot(result=bool)
    def export(self) -> bool:
        if not self.canExport or self._shutting_down or self._worker is not None:
            return False
        owner = object()
        if self._app and not self._app.acquire_task("export", "Excel 내보내기", owner):
            self.set_field_errors({"form": "다른 작업이 종료될 때까지 기다려 주세요."})
            return False
        request = ExportRequest(self._supplier_id, self._output_path)
        worker = self._worker_factory(request)
        self._operation_id += 1
        operation = self._operation_id
        self._task_owner = owner
        self._worker = worker
        self._busy = True
        worker.complete.connect(lambda path: self._on_complete(operation, path))
        worker.error.connect(lambda message: self._on_error(operation, message))
        worker.cancelled.connect(lambda: self._on_cancelled(operation))
        self._emit()
        worker.start()
        return True

    def _current(self, operation: int) -> bool:
        return operation == self._operation_id and not self._shutting_down

    def _retire(self) -> None:
        if self._worker is not None:
            self._retired_workers.append(self._worker)
            self._worker = None
        self._retired_workers = [w for w in self._retired_workers if getattr(w, "isRunning", lambda: False)()]
        if self._retired_workers:
            QTimer.singleShot(25, self._cleanup_retired)

    def _cleanup_retired(self) -> None:
        self._retired_workers = [w for w in self._retired_workers if getattr(w, "isRunning", lambda: False)()]
        if self._retired_workers:
            QTimer.singleShot(25, self._cleanup_retired)
        self._emit()

    def _on_complete(self, operation: int, _path: str) -> None:
        if not self._current(operation): return
        self._busy = False
        self._retire()
        if self._app and self._task_owner is not None: self._app.complete_owned_task(self._task_owner)
        self._task_owner = None
        self.refreshHistory()
        self.validateScope()

    def _on_error(self, operation: int, message: str) -> None:
        if not self._current(operation): return
        self._busy = False
        self._retire()
        safe = sanitize_diagnostic(message)
        self.set_field_errors({"form": safe})
        if self._app and self._task_owner is not None: self._app.fail_owned_task(self._task_owner, safe)
        self._task_owner = None
        self._emit()

    def _on_cancelled(self, operation: int) -> None:
        if not self._current(operation): return
        self._busy = False
        self._retire()
        if self._app and self._task_owner is not None: self._app.cancel_owned_task(self._task_owner)
        self._task_owner = None
        self._emit()

    @Slot()
    def refreshHistory(self) -> None:
        rows = []
        self._exports_dir.mkdir(parents=True, exist_ok=True)
        files = sorted(self._exports_dir.glob("*.xlsx"), key=lambda path: (-path.stat().st_mtime_ns, path.name))[:10]
        for path in files:
            count, outcome = 0, "success"
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                sheet = workbook["products"]
                count = max(0, sheet.max_row - 1)
                workbook.close()
            except Exception:
                outcome = "unreadable"
            rows.append({"fileName": path.name, "path": str(path), "exportedAt": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"), "rowCount": count, "outcome": outcome})
        self._history.resetRows(rows)
        self._emit()

    @Slot(str, str)
    def selectIssue(self, product_id: str, _product_code: str = "") -> None:
        if product_id and self._app:
            self._app.navigate("crawl")

    @Slot()
    def shutdown(self) -> None:
        if self._shutting_down: return
        self._shutting_down = True
        self._operation_id += 1
        if self._app and self._task_owner is not None:
            self._app.cancel_owned_task(self._task_owner, "내보내기 종료됨")
        workers = list(self._retired_workers)
        if self._worker is not None:
            workers.append(self._worker)
            self._worker = None
        self._retired_workers = stop_workers(workers)
        self._busy = False
        self._task_owner = None
        self._emit()
