from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.db.models import Base, Product, ProductOption, Supplier
from app.exporters.excel import WHOLESALE_COLUMNS, export_to_excel


def _setup_db(tmp_path: Path) -> Session:
    engine = create_engine(f"sqlite:///{tmp_path / 'test_export.db'}")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, expire_on_commit=False)()


def _add_test_data(session: Session) -> str:
    supplier = Supplier(name="테스트도매처", base_url="https://test.com")
    session.add(supplier)
    session.flush()

    product = Product(
        supplier_id=supplier.id,
        supplier_name="테스트도매처",
        supplier_product_code="P-001",
        supplier_product_id="1",
        supplier_status="available",
        raw_product_name="테스트 상품",
        origin="국산",
        supply_price=12000,
        main_image_url="https://img/test.jpg",
        detail_content="<p>상세</p>",
        extra_image_urls=["https://img/2.jpg", "https://img/3.jpg"],
    )
    session.add(product)
    session.flush()

    session.add(ProductOption(
        product_id=product.id, option_sku="P-001-1", option_type="combination",
        option_group_1="색상", option_value_1="블랙", option_display_name="블랙",
        option_supply_price=12000, option_main_image_url="https://img/black.jpg",
        option_price_delta=0, option_usable=True, option_position=1,
    ))
    session.add(ProductOption(
        product_id=product.id, option_sku="P-001-2", option_type="combination",
        option_group_1="색상", option_value_1="화이트", option_display_name="화이트",
        option_supply_price=13000, option_main_image_url="https://img/white.jpg",
        option_price_delta=1000, option_usable=True, option_position=2,
    ))
    session.commit()
    return supplier.id


def _single_sheet(output: Path):
    wb = load_workbook(output)
    assert wb.sheetnames == ["상품대장"]  # 단일 시트만 (서버는 첫 시트만 읽음)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    row = {h: ws.cell(row=2, column=i + 1).value for i, h in enumerate(headers)}
    return headers, row


def test_export_uses_single_korean_sheet_with_wholesale_headers(tmp_path: Path) -> None:
    session = _setup_db(tmp_path)
    supplier_id = _add_test_data(session)
    output = tmp_path / "export.xlsx"
    assert export_to_excel(session, supplier_id, output) == output

    headers, _ = _single_sheet(output)
    assert headers == WHOLESALE_COLUMNS
    assert headers[:8] == ["상태", "제품번호", "상품코드", "상품명", "가격", "원산지", "목록이미지1", "상세이미지"]


def test_required_fields_map_from_crawler_product(tmp_path: Path) -> None:
    session = _setup_db(tmp_path)
    supplier_id = _add_test_data(session)
    output = tmp_path / "export.xlsx"
    export_to_excel(session, supplier_id, output)
    _, row = _single_sheet(output)

    assert row["상태"] == "판매중"          # available → 판매중
    assert row["제품번호"] == "1"
    assert row["상품코드"] == "P-001"
    assert row["상품명"] == "테스트 상품"
    assert row["원산지"] == "국산"
    assert row["목록이미지1"] == "https://img/test.jpg"
    assert row["상세이미지"] == "<p>상세</p>"
    # 목록이미지2~5는 extra_image_urls에서
    assert row["목록이미지2"] == "https://img/2.jpg"
    assert row["목록이미지3"] == "https://img/3.jpg"


def test_options_inline_price_and_image_csv_aligned(tmp_path: Path) -> None:
    session = _setup_db(tmp_path)
    supplier_id = _add_test_data(session)
    output = tmp_path / "export.xlsx"
    export_to_excel(session, supplier_id, output)
    _, row = _single_sheet(output)

    # 옵션이 있으면 가격은 옵션별 CSV(옵션값과 1:1 정렬)
    assert row["옵션값"] == "블랙,화이트"
    assert row["가격"] == "12000,13000"
    assert row["옵션이미지"] == "https://img/black.jpg,https://img/white.jpg"


def test_price_falls_back_to_supply_price_without_options(tmp_path: Path) -> None:
    session = _setup_db(tmp_path)
    supplier = Supplier(name="s", base_url="https://x")
    session.add(supplier)
    session.flush()
    session.add(Product(
        supplier_id=supplier.id, supplier_name="s", supplier_product_code="N-1",
        supplier_status="sold_out", raw_product_name="옵션없는상품", supply_price=5000,
    ))
    session.commit()
    output = tmp_path / "export.xlsx"
    export_to_excel(session, supplier.id, output)
    _, row = _single_sheet(output)

    assert row["가격"] == "5000"   # 옵션 없으면 단일 공급가
    assert not row["옵션값"]        # 옵션 없음 (openpyxl은 빈 셀을 None으로 되읽음)
    assert row["상태"] == "품절"    # sold_out → 품절


def test_export_with_no_supplier_id_exports_all(tmp_path: Path) -> None:
    session = _setup_db(tmp_path)
    _add_test_data(session)
    output = tmp_path / "export_all.xlsx"
    export_to_excel(session, None, output)
    wb = load_workbook(output)
    assert wb.active.max_row - 1 == 1
